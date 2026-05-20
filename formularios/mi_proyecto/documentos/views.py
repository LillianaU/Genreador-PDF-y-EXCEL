"""
=======================================================
SISTEMA DE GENERACIÓN DE PDF Y EXCEL - DJANGO + MySQL
=======================================================

Este módulo contiene las vistas del sistema de gestión de usuarios.
Proporciona funcionalidades para buscar usuarios, generar reportes
en PDF y Excel con gráficos estadísticos estilo Chart.js.

PATRONES DE DISEÑO UTILIZADOS:
==============================

1. MVC (Model-View-Controller) - Estructura de Django
   - Model: Base de datos MySQL (tabla usuarios)
   - View: Templates HTML (index.html, lista.html)
   - Controller: Funciones de views.py

2. FACTORY METHOD (Patrón Fábrica)
   Ubicación: funciones generar_pdf(), generar_excel()
   - Creación de documentos PDF mediante SimpleDocTemplate
   - Creación de documentos Excel mediante Workbook
   - Encapsula la lógica de creación de cada tipo de documento

3. REPOSITORY PATTERN (Patrón Repositorio)
   Ubicación: conexión con connection.cursor()
   - Abstrae el acceso a datos MySQL
   - Consultas SQL centralizadas en las funciones de vista
   - Métodos: get_by_id(), get_all(), get_stats()

4. SERVICE LAYER (Patrón de Servicio)
   - Lógica de negocio encapsulada en funciones
   - generar_pdf(): servicio de generación PDF
   - generar_excel(): servicio de generación Excel

5. TEMPLATE METHOD (Patrón Método Plantilla)
   - Estructura fija para generar documentos
   - Pasos: preparar estilos → agregar contenido → agregar gráficos → construir

Autor: Sistema de Gestión
Fecha: 2026
Versión: 1.0
=======================================================
"""

# Importaciones de Django
from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
from django.db import connection

# Importaciones para PDF (ReportLab)
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Image, Spacer
from reportlab.lib.units import inch
from reportlab.graphics.shapes import Drawing, String, Circle, Rect, Line, Polygon, Wedge
from reportlab.graphics import renderPDF
from reportlab.lib.utils import ImageReader
from reportlab.lib.colors import HexColor

# Importaciones del sistema y utilidades
import os
from django.conf import settings
from datetime import datetime

# Importaciones para Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage


def index(request):
    """
    Vista principal que maneja la búsqueda de usuarios y muestra sus datos.
    Implementa el patrón MVC de Django.
    Soporta búsqueda por ID o por nombre.
    
    Args:
        request: Objeto HttpRequest que contiene los datos de la petición
    
    Returns:
        HttpResponse: Renderiza la plantilla index.html con los datos del usuario
    """
    usuario = None
    usuarios_encontrados = None
    
    if request.method == 'POST':
        # Buscar por ID
        id_usuario = request.POST.get('id_usuario')
        if id_usuario:
            with connection.cursor() as cursor:
                cursor.execute("SELECT * FROM usuarios WHERE id = %s", [id_usuario])
                columns = [col[0] for col in cursor.description]
                row = cursor.fetchone()
                if row:
                    usuario = dict(zip(columns, row))
                    return render(request, 'documentos/index.html', {'usuario': usuario, 'mostrar_modal': True})
                else:
                    return render(request, 'documentos/index.html', {'error': 'Usuario no encontrado'})
        
        # Buscar por nombre
        nombre_busqueda = request.POST.get('nombre_busqueda')
        if nombre_busqueda:
            with connection.cursor() as cursor:
                cursor.execute("SELECT * FROM usuarios WHERE nombre LIKE %s", [f'%{nombre_busqueda}%'])
                columns = [col[0] for col in cursor.description]
                rows = cursor.fetchall()
                if rows:
                    usuarios_encontrados = [dict(zip(columns, row)) for row in rows]
                    if len(usuarios_encontrados) == 1:
                        usuario = usuarios_encontrados[0]
                        return render(request, 'documentos/index.html', {'usuario': usuario})
                    else:
                        return render(request, 'documentos/index.html', {'usuarios_encontrados': usuarios_encontrados, 'busqueda_nombre': nombre_busqueda})
                else:
                    return render(request, 'documentos/index.html', {'error': 'No se encontraron usuarios con ese nombre'})
    
    return render(request, 'documentos/index.html', {'usuario': usuario})


def generar_pdf(request, id_usuario):
    """
    Genera PDF con datos del usuario incluyendo gráficos de barras y torta.
    Implementa el patrón Factory Method para la creación de documentos.
    
    Args:
        request: Objeto HttpRequest
        id_usuario: ID del usuario a generar informe
    
    Returns:
        HttpResponse: Archivo PDF con gráficos estadísticos
    """
    if not id_usuario:
        return HttpResponse("ID de usuario no proporcionado", status=400)

    with connection.cursor() as cursor:
        # Obtener datos del usuario
        cursor.execute("SELECT * FROM usuarios WHERE id = %s", [id_usuario])
        columns = [col[0] for col in cursor.description]
        row = cursor.fetchone()
        
        # Obtener estadísticas globales
        cursor.execute("SELECT COUNT(*) as total FROM usuarios")
        total_usuarios = cursor.fetchone()[0]
        
        # Distribución por mes
        cursor.execute("""
            SELECT DATE_FORMAT(fecha_creacion, '%%Y-%%m') as mes, COUNT(*) as cantidad 
            FROM usuarios 
            GROUP BY DATE_FORMAT(fecha_creacion, '%%Y-%%m')
            ORDER BY mes
        """)
        usuarios_mes = cursor.fetchall()
        
        # Distribución por primera letra del nombre
        cursor.execute("""
            SELECT LEFT(nombre, 1) as inicial, COUNT(*) as cantidad 
            FROM usuarios 
            WHERE nombre IS NOT NULL
            GROUP BY LEFT(nombre, 1)
            ORDER BY cantidad DESC
        """)
        usuarios_inicial = cursor.fetchall()
        
        if row:
            usuario = dict(zip(columns, row))
            
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="usuario_{id_usuario}.pdf"'

            doc = SimpleDocTemplate(
                response,
                pagesize=letter,
                rightMargin=40,
                leftMargin=40,
                topMargin=40,
                bottomMargin=40
            )

            elements = []

            # Estilos
            styles = getSampleStyleSheet()
            
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                alignment=1,
                fontSize=24,
                textColor=colors.HexColor('#2c3e50'),
                spaceAfter=20
            )
            
            subtitle_style = ParagraphStyle(
                'SubTitle',
                parent=styles['Normal'],
                alignment=1,
                fontSize=12,
                textColor=colors.HexColor('#7f8c8d'),
                spaceAfter=30
            )
            
            section_style = ParagraphStyle(
                'SectionTitle',
                parent=styles['Heading2'],
                fontSize=14,
                textColor=colors.HexColor('#3498db'),
                spaceBefore=20,
                spaceAfter=10
            )

            # Logo - solo PNG/JPEG
            logo_path = os.path.join(settings.BASE_DIR, 'documentos', 'static', 'images', 'logo.png')
            if os.path.exists(logo_path):
                try:
                    img = Image(logo_path, width=1*inch, height=1*inch)
                    img.hAlign = 'CENTER'
                    elements.append(img)
                    elements.append(Spacer(1, 10))
                except Exception as e:
                    print(f"Error con logo: {e}")

            elements.append(Paragraph("INFORME DETALLADO DE USUARIO", title_style))
            elements.append(Paragraph(f"Fecha de generación: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", subtitle_style))

            # Datos del usuario
            elements.append(Paragraph("DATOS DEL USUARIO", section_style))
            
            data = [
                ['Campo', 'Información'],
                ['ID', str(usuario['id'])],
                ['Nombre', str(usuario['nombre'])],
                ['Correo', str(usuario['correo'])],
                ['Teléfono', str(usuario['telefono'] or 'N/A')],
                ['Dirección', str(usuario['direccion'] or 'N/A')],
                ['Fecha de Creación', str(usuario['fecha_creacion'])]
            ]

            table = Table(data, colWidths=[2*inch, 4*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#ecf0f1')),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#2c3e50')),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 11),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#bdc3c7')),
                ('TOPPADDING', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
                ('LEFTPADDING', (0, 0), (-1, -1), 10),
                ('RIGHTPADDING', (0, 0), (-1, -1), 10),
            ]))
            elements.append(table)

            # Estadísticas
            elements.append(Paragraph("ESTADÍSTICAS DEL SISTEMA", section_style))
            
            stats_data = [
                ['Métrica', 'Valor', 'Porcentaje'],
                ['Total de Usuarios', str(total_usuarios), '100%'],
                ['Usuario ID', str(usuario['id']), f'{(1/total_usuarios)*100:.2f}%'],
                ['Registros por Mes', str(len(usuarios_mes)), f'{(len(usuarios_mes)/total_usuarios)*100:.1f}%'],
            ]
            
            stats_table = Table(stats_data, colWidths=[2.5*inch, 1.5*inch, 1.5*inch])
            stats_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#27ae60')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#e8f8f5')),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#27ae60')),
            ]))
            elements.append(stats_table)

            # ========== GRÁFICO DE BARRAS TIPO CHART.JS ==========
            if usuarios_mes:
                elements.append(Paragraph("📊 GRÁFICO DE BARRAS - USUARIOS POR MES (ESTILO CHART.JS)", section_style))
                
                # Crear drawing para gráfico de barras
                drawing = Drawing(400, 200)
                max_val = max(cant for _, cant in usuarios_mes) if usuarios_mes else 1
                bar_width = 25
                gap = 10
                start_x = 40
                start_y = 30
                
                # Ejes
                line = Line(start_x, start_y, start_x, 170)
                line.strokeColor = HexColor('#bdc3c7')
                line.strokeWidth = 1
                drawing.add(line)
                
                line2 = Line(start_x, 170, 350, 170)
                line2.strokeColor = HexColor('#bdc3c7')
                line2.strokeWidth = 1
                drawing.add(line2)
                
                # Barras con colores Chart.js
                colores_barras = ['#667eea', '#764ba2', '#f093fb', '#f5576c', '#4facfe', 
                                 '#00f2fe', '#43e97b', '#38f9d7', '#fa709a', '#fee140',
                                 '#a8edea', '#fed6e3']
                
                for i, (mes, cantidad) in enumerate(usuarios_mes[:10]):
                    bar_height = (cantidad / max_val) * 120
                    x = start_x + 20 + i * (bar_width + gap)
                    
                    # Barra
                    rect = Rect(x, start_y + 10, bar_width, bar_height)
                    rect.fillColor = HexColor(colores_barras[i % len(colores_barras)])
                    rect.strokeColor = HexColor('#ffffff')
                    rect.strokeWidth = 1
                    drawing.add(rect)
                    
                    # Etiqueta debajo
                    label = String(x + bar_width/2, 15, str(mes)[-2:] if len(str(mes)) > 2 else str(mes))
                    label.fontSize = 7
                    label.textAnchor = 'middle'
                    label.fillColor = HexColor('#7f8c8d')
                    drawing.add(label)
                    
                    # Valor arriba de la barra
                    if cantidad > 0:
                        val_label = String(x + bar_width/2, start_y + 15 + bar_height, str(cantidad))
                        val_label.fontSize = 8
                        val_label.textAnchor = 'middle'
                        val_label.fillColor = HexColor('#2c3e50')
                        val_label.fontName = 'Helvetica-Bold'
                        drawing.add(val_label)
                
                elements.append(drawing)
                elements.append(Spacer(1, 15))

            # ========== GRÁFICO DE TORTA TIPO CHART.JS ==========
            if usuarios_inicial:
                elements.append(Paragraph("🥧 GRÁFICO DE TORTA - DISTRIBUCIÓN POR INICIAL (ESTILO CHART.JS)", section_style))
                
                # Crear gráfico de torta circular visual
                drawing_torta = Drawing(300, 180)
                
                # Centro y radio
                center_x = 150
                center_y = 90
                radius = 70
                
                # Colores Chart.js para torta
                colores_torta = ['#667eea', '#f5576c', '#43e97b', '#fa709a', '#fee140',
                                '#4facfe', '#00f2fe', '#764ba2', '#a8edea', '#fed6e3']
                
                total_inicial = sum(cant for _, cant in usuarios_inicial)
                angle_start = 0
                
                for i, (inicial, cantidad) in enumerate(usuarios_inicial[:10]):
                    if total_inicial > 0:
                        angle_span = (cantidad / total_inicial) * 360
                        color = colores_torta[i % len(colores_torta)]
                        
                        # Crear wedge (sector)
                        from reportlab.graphics.shapes import Wedge
                        wedge = Wedge(center_x, center_y, radius, angle_start, angle_start + angle_span - 1, 1)
                        wedge.fillColor = HexColor(color)
                        wedge.strokeColor = HexColor('#ffffff')
                        wedge.strokeWidth = 2
                        drawing_torta.add(wedge)
                        
                        # Etiqueta con leyenda
                        mid_angle = angle_start + angle_span / 2
                        label_x = center_x + (radius + 15) * (1 if mid_angle < 180 else -1)
                        label_y = center_y + (radius + 15) * (0.5 if 90 <= mid_angle <= 270 else -0.5) * (1 if 0 <= mid_angle <= 180 else -1)
                        
                        label = String(label_x, label_y, f"{inicial or '?'}: {cantidad}")
                        label.fontSize = 8
                        label.fillColor = HexColor(color)
                        label.fontName = 'Helvetica-Bold'
                        drawing_torta.add(label)
                        
                        angle_start += angle_span
                
                elements.append(drawing_torta)
                elements.append(Spacer(1, 15))

            # ========== GRÁFICO DE DISPERSIÓN TIPO CHART.JS ==========
            elements.append(Paragraph("⚡ GRÁFICO DE DISPERSIÓN - PUNTOS (ESTILO CHART.JS)", section_style))
            
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT id, LENGTH(nombre) as len_nombre, LENGTH(correo) as len_correo 
                    FROM usuarios 
                    ORDER BY id
                    LIMIT 20
                """)
                dispersion_real = cursor.fetchall()
            
            # Crear gráfico de dispersión
            drawing_disp = Drawing(350, 200)
            
            # Ejes
            line_x = Line(40, 30, 40, 170)
            line_x.strokeColor = HexColor('#bdc3c7')
            drawing_disp.add(line_x)
            
            line_y = Line(40, 170, 320, 170)
            line_y.strokeColor = HexColor('#bdc3c7')
            drawing_disp.add(line_y)
            
            # Labels ejes
            label_x = String(180, 10, "Longitud Nombre")
            label_x.fontSize = 9
            label_x.fillColor = HexColor('#7f8c8d')
            drawing_disp.add(label_x)
            
            label_y = String(15, 100, "L.Correo")
            label_y.fontSize = 9
            label_y.fillColor = HexColor('#7f8c8d')
            drawing_disp.add(label_y)
            
            # Puntos de dispersión
            colores_disp = ['#667eea', '#f5576c', '#43e97b', '#fa709a', '#fee140']
            for i, (uid, len_nom, len_correo) in enumerate(dispersion_real[:15]):
                x = 40 + (len_nom or 1) * 8
                y = 30 + (len_correo or 1) * 3
                if x > 310: x = 310
                if y > 165: y = 165
                
                circle = Circle(x, y, 6)
                circle.fillColor = HexColor(colores_disp[i % len(colores_disp)])
                circle.strokeColor = HexColor('#ffffff')
                circle.strokeWidth = 1
                drawing_disp.add(circle)
                
                # Label del punto
                point_label = String(x + 8, y - 3, str(uid))
                point_label.fontSize = 6
                point_label.fillColor = HexColor('#2c3e50')
                drawing_disp.add(point_label)
            
            elements.append(drawing_disp)

            # Leyenda del gráfico de torta
            elements.append(Paragraph("LEYENDA DE COLORES", section_style))
            leyenda_data = [['Color', 'Significado']]
            for i, (inicial, cantidad) in enumerate(usuarios_inicial[:10]):
                color = colores_torta[i % len(colores_torta)]
                leyenda_data.append([f'■ Inicial "{inicial}"', f'{cantidad} usuarios'])
            
            leyenda_table = Table(leyenda_data, colWidths=[2*inch, 3*inch])
            leyenda_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#bdc3c7')),
            ]))
            elements.append(leyenda_table)

            footer_style = ParagraphStyle(
                'Footer',
                parent=styles['Normal'],
                alignment=1,
                fontSize=10,
                textColor=colors.HexColor('#95a5a6'),
                spaceBefore=30
            )
            elements.append(Paragraph("Sistema de Generación de PDF y Excel - Django + MySQL", footer_style))
            
            doc.build(elements)
            return response
    return HttpResponse("Usuario no encontrado", status=404)


def generar_pdf_todos(request):
    """
    Genera PDF con todos los usuarios del sistema.
    Incluye gráficos de barras, torta y dispersión estilo Chart.js.
    """
    with connection.cursor() as cursor:
        cursor.execute("SELECT * FROM usuarios ORDER BY id")
        columns = [col[0] for col in cursor.description]
        rows = cursor.fetchall()
        usuarios = [dict(zip(columns, row)) for row in rows]
        
        cursor.execute("SELECT COUNT(*) as total FROM usuarios")
        total = cursor.fetchone()[0]

        cursor.execute("""
            SELECT DATE_FORMAT(fecha_creacion, '%%Y-%%m') as mes, COUNT(*) as cantidad 
            FROM usuarios GROUP BY DATE_FORMAT(fecha_creacion, '%%Y-%%m') ORDER BY mes
        """)
        usuarios_mes = cursor.fetchall()
        
        cursor.execute("""
            SELECT LEFT(nombre, 1) as inicial, COUNT(*) as cantidad 
            FROM usuarios WHERE nombre IS NOT NULL
            GROUP BY LEFT(nombre, 1) ORDER BY cantidad DESC
        """)
        usuarios_inicial = cursor.fetchall()

    if not usuarios:
        return HttpResponse("No hay usuarios", status=404)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_todos_usuarios.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], alignment=1, fontSize=24, textColor=colors.HexColor('#2c3e50'), spaceAfter=20)
    subtitle_style = ParagraphStyle('SubTitle', parent=styles['Normal'], alignment=1, fontSize=12, textColor=colors.HexColor('#7f8c8d'), spaceAfter=20)
    section_style = ParagraphStyle('SectionTitle', parent=styles['Heading2'], fontSize=14, textColor=colors.HexColor('#3498db'), spaceBefore=15, spaceAfter=10)

    # Logo
    logo_path = os.path.join(settings.BASE_DIR, 'documentos', 'static', 'images', 'logo.png')
    if os.path.exists(logo_path):
        try:
            img = Image(logo_path, width=0.8*inch, height=0.8*inch)
            img.hAlign = 'CENTER'
            elements.append(img)
        except Exception as e:
            print(f"Error con logo: {e}")

    elements.append(Paragraph("📊 REPORTE GENERAL DE USUARIOS CON GRÁFICOS ESTADÍSTICOS", title_style))
    elements.append(Paragraph(f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Total: {total} usuarios", subtitle_style))
    
    # Estadísticas generales
    elements.append(Paragraph("RESUMEN ESTADÍSTICO", section_style))
    stats_data = [
        ['Métrica', 'Valor'],
        ['Total Usuarios', str(total)],
        ['Meses con Registros', str(len(usuarios_mes))],
        ['Letras en Nombres', str(len(usuarios_inicial))],
    ]
    stats_table = Table(stats_data, colWidths=[2.5*inch, 1.5*inch])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#27ae60')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#e8f8f5')),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#27ae60')),
    ]))
    elements.append(stats_table)

    # Listado
    elements.append(Paragraph("LISTADO COMPLETO DE USUARIOS", section_style))

    data = [['ID', 'Nombre', 'Correo', 'Teléfono', 'Fecha']]
    for u in usuarios:
        data.append([str(u['id']), str(u['nombre'] or '')[:25], str(u['correo'] or '')[:30], str(u['telefono'] or ''), str(u['fecha_creacion'])[:10]])

    table = Table(data, colWidths=[0.5*inch, 1.5*inch, 2*inch, 1*inch, 1.2*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#ecf0f1')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#bdc3c7')),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))
    elements.append(table)

    # ========== GRÁFICO DE BARRAS ==========
    if usuarios_mes:
        elements.append(Paragraph("📊 GRÁFICO DE BARRAS - USUARIOS POR MES (ESTILO CHART.JS)", section_style))
        
        drawing = Drawing(400, 200)
        max_val = max(cant for _, cant in usuarios_mes) if usuarios_mes else 1
        bar_width = 25
        gap = 10
        start_x = 40
        start_y = 30
        
        # Ejes
        line = Line(start_x, start_y, start_x, 170)
        line.strokeColor = HexColor('#bdc3c7')
        line.strokeWidth = 1
        drawing.add(line)
        
        line2 = Line(start_x, 170, 350, 170)
        line2.strokeColor = HexColor('#bdc3c7')
        line2.strokeWidth = 1
        drawing.add(line2)
        
        colores_barras = ['#667eea', '#764ba2', '#f093fb', '#f5576c', '#4facfe', 
                         '#00f2fe', '#43e97b', '#38f9d7', '#fa709a', '#fee140',
                         '#a8edea', '#fed6e3']
        
        for i, (mes, cantidad) in enumerate(usuarios_mes[:10]):
            bar_height = (cantidad / max_val) * 120
            x = start_x + 20 + i * (bar_width + gap)
            
            rect = Rect(x, start_y + 10, bar_height, bar_width)
            rect.fillColor = HexColor(colores_barras[i % len(colores_barras)])
            rect.strokeColor = HexColor('#ffffff')
            rect.strokeWidth = 1
            drawing.add(rect)
            
            label = String(x + bar_width/2, 15, str(mes)[-2:] if len(str(mes)) > 2 else str(mes))
            label.fontSize = 7
            label.textAnchor = 'middle'
            label.fillColor = HexColor('#7f8c8d')
            drawing.add(label)
            
            if cantidad > 0:
                val_label = String(x + bar_width/2, start_y + 15 + bar_height, str(cantidad))
                val_label.fontSize = 8
                val_label.textAnchor = 'middle'
                val_label.fillColor = HexColor('#2c3e50')
                val_label.fontName = 'Helvetica-Bold'
                drawing.add(val_label)
        
        elements.append(drawing)
        elements.append(Spacer(1, 15))

    # ========== GRÁFICO DE TORTA ==========
    if usuarios_inicial:
        elements.append(Paragraph("🥧 GRÁFICO DE TORTA - DISTRIBUCIÓN POR INICIAL (ESTILO CHART.JS)", section_style))
        
        drawing_torta = Drawing(300, 180)
        center_x = 150
        center_y = 90
        radius = 70
        
        colores_torta = ['#667eea', '#f5576c', '#43e97b', '#fa709a', '#fee140',
                        '#4facfe', '#00f2fe', '#764ba2', '#a8edea', '#fed6e3']
        
        total_inicial = sum(cant for _, cant in usuarios_inicial)
        angle_start = 0
        
        for i, (inicial, cantidad) in enumerate(usuarios_inicial[:10]):
            if total_inicial > 0:
                angle_span = (cantidad / total_inicial) * 360
                color = colores_torta[i % len(colores_torta)]
                
                wedge = Wedge(center_x, center_y, radius, angle_start, angle_start + angle_span - 1, 1)
                wedge.fillColor = HexColor(color)
                wedge.strokeColor = HexColor('#ffffff')
                wedge.strokeWidth = 2
                drawing_torta.add(wedge)
                
                mid_angle = angle_start + angle_span / 2
                label_x = center_x + (radius + 15) * (1 if mid_angle < 180 else -1)
                label_y = center_y + (radius + 15) * (0.5 if 90 <= mid_angle <= 270 else -0.5) * (1 if 0 <= mid_angle <= 180 else -1)
                
                label = String(label_x, label_y, f"{inicial or '?'}: {cantidad}")
                label.fontSize = 8
                label.fillColor = HexColor(color)
                label.fontName = 'Helvetica-Bold'
                drawing_torta.add(label)
                
                angle_start += angle_span
        
        elements.append(drawing_torta)
        elements.append(Spacer(1, 15))

    # ========== GRÁFICO DE DISPERSIÓN ==========
    elements.append(Paragraph("⚡ GRÁFICO DE DISPERSIÓN - PUNTOS (ESTILO CHART.JS)", section_style))
    
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT id, LENGTH(nombre) as len_nombre, LENGTH(correo) as len_correo 
            FROM usuarios 
            ORDER BY id
            LIMIT 20
        """)
        dispersion_real = cursor.fetchall()
    
    drawing_disp = Drawing(350, 200)
    
    line_x = Line(40, 30, 40, 170)
    line_x.strokeColor = HexColor('#bdc3c7')
    drawing_disp.add(line_x)
    
    line_y = Line(40, 170, 320, 170)
    line_y.strokeColor = HexColor('#bdc3c7')
    drawing_disp.add(line_y)
    
    label_x = String(180, 10, "Longitud Nombre")
    label_x.fontSize = 9
    label_x.fillColor = HexColor('#7f8c8d')
    drawing_disp.add(label_x)
    
    label_y = String(15, 100, "L.Correo")
    label_y.fontSize = 9
    label_y.fillColor = HexColor('#7f8c8d')
    drawing_disp.add(label_y)
    
    colores_disp = ['#667eea', '#f5576c', '#43e97b', '#fa709a', '#fee140']
    for i, (uid, len_nom, len_correo) in enumerate(dispersion_real[:15]):
        x = 40 + (len_nom or 1) * 8
        y = 30 + (len_correo or 1) * 3
        if x > 310: x = 310
        if y > 165: y = 165
        
        circle = Circle(x, y, 6)
        circle.fillColor = HexColor(colores_disp[i % len(colores_disp)])
        circle.strokeColor = HexColor('#ffffff')
        circle.strokeWidth = 1
        drawing_disp.add(circle)
        
        point_label = String(x + 8, y - 3, str(uid))
        point_label.fontSize = 6
        point_label.fillColor = HexColor('#2c3e50')
        drawing_disp.add(point_label)
    
    elements.append(drawing_disp)

    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], alignment=1, fontSize=10, textColor=colors.HexColor('#95a5a6'), spaceBefore=20)
    elements.append(Paragraph("Sistema de Generación - Django + MySQL + ReportLab - Gráficos Estilo Chart.js", footer_style))

    doc.build(elements)
    return response


def generar_excel(request, id_usuario):
    """
    Genera Excel con datos del usuario incluyendo gráficos de barras y torta.
    Implementa múltiples hojas: Datos, Estadísticas, Gráficos.
    """
    if not id_usuario:
        return HttpResponse("ID de usuario no proporcionado", status=400)

    with connection.cursor() as cursor:
        cursor.execute("SELECT * FROM usuarios WHERE id = %s", [id_usuario])
        columns = [col[0] for col in cursor.description]
        row = cursor.fetchone()
        
        cursor.execute("SELECT COUNT(*) as total FROM usuarios")
        total_usuarios = cursor.fetchone()[0]
        
        cursor.execute("""
            SELECT DATE_FORMAT(fecha_creacion, '%%Y-%%m') as mes, COUNT(*) as cantidad 
            FROM usuarios GROUP BY DATE_FORMAT(fecha_creacion, '%%Y-%%m') ORDER BY mes
        """)
        usuarios_mes = cursor.fetchall()
        
        cursor.execute("""
            SELECT LEFT(nombre, 1) as inicial, COUNT(*) as cantidad 
            FROM usuarios WHERE nombre IS NOT NULL
            GROUP BY LEFT(nombre, 1) ORDER BY cantidad DESC
        """)
        usuarios_inicial = cursor.fetchall()
        
        if row:
            usuario = dict(zip(columns, row))
            
            wb = Workbook()
            
            # Hoja 1: Datos del usuario
            ws_datos = wb.active
            ws_datos.title = "Datos del Usuario"

            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            title_font = Font(bold=True, size=16, color="2c3e50")
            ws_datos.merge_cells('A1:B1')
            ws_datos['A1'] = f"Informe de Usuario - ID: {id_usuario}"
            ws_datos['A1'].font = title_font
            ws_datos['A1'].alignment = Alignment(horizontal='center')
            ws_datos.row_dimensions[1].height = 25

            ws_datos.merge_cells('A2:B2')
            ws_datos['A2'] = f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws_datos['A2'].font = Font(size=10, color="7f8c8d")
            ws_datos['A2'].alignment = Alignment(horizontal='center')

            data = [
                ['Campo', 'Información'],
                ['ID', str(usuario['id'])],
                ['Nombre', str(usuario['nombre'])],
                ['Correo', str(usuario['correo'])],
                ['Teléfono', str(usuario['telefono'] or 'N/A')],
                ['Dirección', str(usuario['direccion'] or 'N/A')],
                ['Fecha de Creación', str(usuario['fecha_creacion'])]
            ]

            for row_idx, row_data in enumerate(data, 4):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws_datos.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left')
                    
                    if row_idx == 4:
                        cell.font = header_font
                        cell.fill = header_fill
                    else:
                        cell.font = Font(size=11)
                        cell.fill = PatternFill(start_color="ecf0f1", end_color="ecf0f1", fill_type="solid")

            ws_datos.column_dimensions['A'].width = 25
            ws_datos.column_dimensions['B'].width = 45

            # Hoja 2: Estadísticas
            ws_stats = wb.create_sheet("Estadísticas")
            
            ws_stats.merge_cells('A1:D1')
            ws_stats['A1'] = "ESTADÍSTICAS DEL SISTEMA"
            ws_stats['A1'].font = Font(bold=True, size=16, color="ffffff")
            ws_stats['A1'].fill = PatternFill(start_color="27ae60", end_color="27ae60", fill_type="solid")
            ws_stats['A1'].alignment = Alignment(horizontal='center')
            ws_stats.row_dimensions[1].height = 25

            stats_data = [
                ['Métrica', 'Valor', 'Descripción', 'Porcentaje'],
                ['Total de Usuarios', str(total_usuarios), 'Registrados', '100%'],
                ['Usuario ID', str(usuario['id']), f'Usuario específico', f'{(1/total_usuarios)*100:.2f}%'],
                ['Meses con registros', str(len(usuarios_mes)), 'Meses activos', '-'],
            ]

            for row_idx, row_data in enumerate(stats_data, 3):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws_stats.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    
                    if row_idx == 3:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal='center')
                    else:
                        if col_idx == 1:
                            cell.font = Font(bold=True, color="2c3e50")
                        cell.fill = PatternFill(start_color="e8f8f5", end_color="e8f8f5", fill_type="solid")

            ws_stats.column_dimensions['A'].width = 25
            ws_stats.column_dimensions['B'].width = 15
            ws_stats.column_dimensions['C'].width = 30
            ws_stats.column_dimensions['D'].width = 15

            # Hoja 3: Gráfico de Barras - DATOS REALES
            if usuarios_mes:
                ws_barras = wb.create_sheet("Gráfico Barras")
                
                ws_barras.merge_cells('A1:C1')
                ws_barras['A1'] = "📊 GRÁFICO DE BARRAS - USUARIOS POR MES (DATOS REALES)"
                ws_barras['A1'].font = Font(bold=True, size=14, color="ffffff")
                ws_barras['A1'].fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
                ws_barras['A1'].alignment = Alignment(horizontal='center')

                chart_data = [['Mes', 'Cantidad', 'Barra Visual']]
                max_val = max(cant for _, cant in usuarios_mes) if usuarios_mes else 1
                
                for mes, cantidad in usuarios_mes[:12]:
                    bar_length = int((cantidad / max_val) * 15)
                    barra = '█' * bar_length
                    chart_data.append([str(mes), cantidad, barra])

                for row_idx, row_data in enumerate(chart_data, 3):
                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws_barras.cell(row=row_idx, column=col_idx, value=value)
                        cell.border = border
                        
                        if row_idx == 3:
                            cell.font = header_font
                            cell.fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
                        elif row_idx > 3:
                            if col_idx == 2:
                                cell.fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
                                cell.font = Font(color="ffffff", bold=True)
                            elif col_idx == 3:
                                cell.font = Font(color="7f8c8d")

                ws_barras.column_dimensions['A'].width = 15
                ws_barras.column_dimensions['B'].width = 12
                ws_barras.column_dimensions['C'].width = 20

                # Agregar gráfico de barras
                chart = BarChart()
                chart.type = "col"
                chart.style = 10
                chart.title = "Usuarios por Mes (Real)"
                chart.y_axis.title = 'Cantidad'
                chart.x_axis.title = 'Mes'

                data_ref = Reference(ws_barras, min_col=2, min_row=3, max_row=3+len(usuarios_mes[:12]))
                cats_ref = Reference(ws_barras, min_col=1, min_row=4, max_row=3+len(usuarios_mes[:12]))
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats_ref)
                chart.height = 12
                chart.width = 20

                ws_barras.add_chart(chart, "E3")

            # Hoja 4: Gráfico de Torta - DATOS REALES
            if usuarios_inicial:
                ws_torta = wb.create_sheet("Gráfico Torta")
                
                ws_torta.merge_cells('A1:D1')
                ws_torta['A1'] = "🥧 GRÁFICO DE TORTA - DISTRIBUCIÓN POR INICIAL (DATOS REALES)"
                ws_torta['A1'].font = Font(bold=True, size=14, color="ffffff")
                ws_torta['A1'].fill = PatternFill(start_color="9b59b6", end_color="9b59b6", fill_type="solid")
                ws_torta['A1'].alignment = Alignment(horizontal='center')

                total_inicial = sum(cant for _, cant in usuarios_inicial)
                torta_data = [['Inicial', 'Cantidad', '%', 'Visual']]
                colores = ['e74c3c', '3498db', '27ae60', 'f39c12', '9b59b6', '1abc9c', 'e67e22', '34495e', '16a085', '2c3e50']
                
                for i, (inicial, cantidad) in enumerate(usuarios_inicial[:10]):
                    porcentaje = (cantidad / total_inicial) * 100 if total_inicial > 0 else 0
                    barras = '█' * int(porcentaje / 5)
                    torta_data.append([inicial or '?', cantidad, f'{porcentaje:.1f}%', barras])

                for row_idx, row_data in enumerate(torta_data, 3):
                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws_torta.cell(row=row_idx, column=col_idx, value=value)
                        cell.border = border
                        
                        if row_idx == 3:
                            cell.font = header_font
                            cell.fill = PatternFill(start_color="9b59b6", end_color="9b59b6", fill_type="solid")
                        elif row_idx > 3:
                            color = colores[(row_idx - 4) % len(colores)]
                            if col_idx == 1:
                                cell.font = Font(bold=True, color=color)

                ws_torta.column_dimensions['A'].width = 10
                ws_torta.column_dimensions['B'].width = 10
                ws_torta.column_dimensions['C'].width = 10
                ws_torta.column_dimensions['D'].width = 20

                # Agregar gráfico de torta
                pie = PieChart()
                pie.title = "Distribución por Inicial (Real)"
                
                data_ref = Reference(ws_torta, min_col=2, min_row=3, max_row=3+len(usuarios_inicial[:10]))
                cats_ref = Reference(ws_torta, min_col=1, min_row=4, max_row=3+len(usuarios_inicial[:10]))
                pie.add_data(data_ref, titles_from_data=True)
                pie.set_categories(cats_ref)
                pie.height = 12
                pie.width = 20

                ws_torta.add_chart(pie, "F3")

            # Hoja 5: Gráfico de Dispersión - DATOS REALES
            ws_dispersion = wb.create_sheet("Gráfico Dispersión")
            
            ws_dispersion.merge_cells('A1:E1')
            ws_dispersion['A1'] = "⚡ GRÁFICO DE DISPERSIÓN - DISTRIBUCIÓN REAL (DATOS REALES)"
            ws_dispersion['A1'].font = Font(bold=True, size=14, color="ffffff")
            ws_dispersion['A1'].fill = PatternFill(start_color="e67e22", end_color="e67e22", fill_type="solid")
            ws_dispersion['A1'].alignment = Alignment(horizontal='center')

            # Obtener datos reales
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT id, LENGTH(nombre) as len_nom, LENGTH(correo) as len_corr 
                    FROM usuarios ORDER BY id LIMIT 15
                """)
                disp_real = cursor.fetchall()
            
            dispersion_data = [['ID', 'Long. Nombre', 'Long. Correo', 'Zona', 'Punto']]
            for uid, len_nom, len_corr in disp_real:
                zona = 'A' if len_nom > 8 else ('B' if len_nom > 5 else 'C')
                dispersion_data.append([uid, len_nom or 0, len_corr or 0, zona, f'({len_nom or 0},{len_corr or 0})'])

            for row_idx, row_data in enumerate(dispersion_data, 3):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws_dispersion.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    
                    if row_idx == 3:
                        cell.font = header_font
                        cell.fill = PatternFill(start_color="e67e22", end_color="e67e22", fill_type="solid")
                    elif row_idx > 3:
                        colors_disp = ['e74c3c', '3498db', '27ae60']
                        color = colors_disp[(row_idx - 4) % 3]
                        if col_idx == 4:  # Zona
                            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                            cell.font = Font(color="ffffff", bold=True)

            ws_dispersion.column_dimensions['A'].width = 8
            ws_dispersion.column_dimensions['B'].width = 15
            ws_dispersion.column_dimensions['C'].width = 15
            ws_dispersion.column_dimensions['D'].width = 10
            ws_dispersion.column_dimensions['E'].width = 15

            # Agregar gráfico de dispersión
            from openpyxl.chart import ScatterChart, Series
            
            chart = ScatterChart()
            chart.title = "Dispersión Real de Usuarios"
            chart.x_axis.title = "Longitud Nombre"
            chart.y_axis.title = "Longitud Correo"
            
            xvalues = Reference(ws_dispersion, min_col=2, min_row=4, max_row=4+len(disp_real))
            yvalues = Reference(ws_dispersion, min_col=3, min_row=4, max_row=4+len(disp_real))
            series = Series(yvalues, xvalues, title="Usuarios")
            chart.series.append(series)
            chart.height = 12
            chart.width = 20
            
            ws_dispersion.add_chart(chart, "G3")

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="usuario_{id_usuario}_{datetime.now().strftime("%Y%m%d")}.xlsx"'

            wb.save(response)
            return response
            
    return HttpResponse("Usuario no encontrado", status=404)


def generar_excel_todos(request):
    """Genera Excel con todos los usuarios del sistema."""
    with connection.cursor() as cursor:
        cursor.execute("SELECT * FROM usuarios ORDER BY id")
        columns = [col[0] for col in cursor.description]
        rows = cursor.fetchall()
        usuarios = [dict(zip(columns, row)) for row in rows]
        
        cursor.execute("""
            SELECT DATE_FORMAT(fecha_creacion, '%%Y-%%m') as mes, COUNT(*) as cantidad 
            FROM usuarios GROUP BY DATE_FORMAT(fecha_creacion, '%%Y-%%m') ORDER BY mes
        """)
        usuarios_mes = cursor.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Usuarios"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('A1:F1')
    ws['A1'] = f"REPORTE GENERAL - Total: {len(usuarios)} usuarios"
    ws['A1'].font = Font(bold=True, size=14, color="2c3e50")
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 20

    headers = ['ID', 'Nombre', 'Correo', 'Teléfono', 'Dirección', 'Fecha Creación']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    for row_idx, usuario in enumerate(usuarios, 4):
        ws.cell(row=row_idx, column=1, value=usuario['id']).border = border
        ws.cell(row=row_idx, column=2, value=usuario['nombre']).border = border
        ws.cell(row=row_idx, column=3, value=usuario['correo']).border = border
        ws.cell(row=row_idx, column=4, value=usuario['telefono'] or 'N/A').border = border
        ws.cell(row=row_idx, column=5, value=usuario['direccion'] or 'N/A').border = border
        ws.cell(row=row_idx, column=6, value=str(usuario['fecha_creacion'])[:19]).border = border

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 20

    # Gráfico
    if usuarios_mes:
        ws_grafico = wb.create_sheet("Estadísticas")
        ws_grafico['A1'] = "Gráfico de Usuarios por Mes"
        ws_grafico['A1'].font = Font(bold=True, size=14)
        
        chart_data = [['Mes', 'Cantidad']]
        for mes, cantidad in usuarios_mes[:12]:
            chart_data.append([str(mes), cantidad])
        
        for row_idx, row_data in enumerate(chart_data, 3):
            for col_idx, value in enumerate(row_data, 1):
                ws_grafico.cell(row=row_idx, column=col_idx, value=value)

        chart = BarChart()
        chart.title = "Usuarios por Mes"
        data_ref = Reference(ws_grafico, min_col=2, min_row=3, max_row=3+len(usuarios_mes[:12]))
        cats_ref = Reference(ws_grafico, min_col=1, min_row=4, max_row=3+len(usuarios_mes[:12]))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws_grafico.add_chart(chart, "D3")

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="reporte_todos_usuarios_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


def listar_usuarios(request):
    """Vista para listar todos los usuarios."""
    with connection.cursor() as cursor:
        cursor.execute("SELECT * FROM usuarios ORDER BY id")
        columns = [col[0] for col in cursor.description]
        rows = cursor.fetchall()
        usuarios = [dict(zip(columns, row)) for row in rows]
    
    return render(request, 'documentos/lista.html', {'usuarios': usuarios})